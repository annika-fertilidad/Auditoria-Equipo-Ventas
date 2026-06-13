#!/usr/bin/env python3
"""
Acumula el export diario de Atom en un histórico para el dashboard PXI.

Cada descarga de Atom (datos.xlsx) contiene SOLO el día. Para que la semana se
acumule —y para que un agente en su día de descanso no desaparezca— este script
fusiona las filas nuevas dentro de datos_historico.csv, eliminando duplicados.

La de-duplicación usa la fila completa como clave, así que volver a exportar el
mismo día NO duplica nada; solo se agregan mensajes/conversaciones nuevos.

Uso:
    python scripts/merge_history.py            # fusiona datos.xlsx -> datos_historico.csv
    python scripts/merge_history.py --daily otro.xlsx --hist datos_historico.csv
"""
import argparse, csv, os, re, sys
from datetime import date, datetime, timedelta

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl (pip install openpyxl)")

# Atom escribe fechas relativas ("hoy a las 6:56 pm", "ayer a las 7:01 pm") para
# los mensajes recientes. Eso es tóxico para un histórico: "ayer" cambia de
# significado según el día de descarga y rompería la de-duplicación. Las
# convertimos a fecha absoluta usando la fecha del export (la del archivo / la
# corrida del Action) ANTES de fusionar, para que queden fijas y comparables.
_REL = re.compile(r"^\s*(hoy|ayer)\s+a\s+las\s+(.+)$", re.I)


def resolve_relative(hora, export_date):
    m = _REL.match(str(hora or ""))
    if not m:
        return str(hora or "").strip()
    when = m.group(1).lower()
    rest = m.group(2).strip()
    d = export_date if when == "hoy" else export_date - timedelta(days=1)
    return f"{d.day:02d}/{d.month:02d}/{d.year} a las {rest}"

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

# Orden canónico de columnas (igual que la plantilla del dashboard).
HEADERS = ["num_conversacion", "cliente_csv", "contacto", "fecha_inicio_gestion",
           "canal", "agente", "tipificacion", "es_venta", "tipo", "direccion",
           "remitente", "contenido", "hora", "url"]


def norm_header(h):
    return str(h or "").strip().lower()


def read_xlsx(path, export_date):
    """Devuelve (lista de filas como dicts) desde la hoja Historial.

    Resuelve las fechas relativas ("hoy/ayer a las ...") de la columna `hora`
    a fecha absoluta usando export_date, para que queden fijas y comparables.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    name = next((n for n in wb.sheetnames if norm_header(n) == "historial"), wb.sheetnames[0])
    ws = wb[name]
    it = ws.iter_rows(values_only=True)
    try:
        header = [norm_header(h) for h in next(it)]
    except StopIteration:
        return []
    idx = {h: i for i, h in enumerate(header)}
    rows = []
    for r in it:
        if r is None:
            continue
        row = {}
        for col in HEADERS:
            i = idx.get(col)
            v = r[i] if (i is not None and i < len(r)) else ""
            row[col] = "" if v is None else str(v).strip()
        row["hora"] = resolve_relative(row.get("hora"), export_date)
        if any(row.values()):
            rows.append(row)
    return rows


def read_csv(path):
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return [{k: (row.get(k) or "") for k in HEADERS} for row in csv.DictReader(f)]


def row_key(row):
    return tuple(row.get(c, "") for c in HEADERS)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--daily", default=os.path.join(ROOT, "datos.xlsx"))
    ap.add_argument("--hist", default=os.path.join(ROOT, "datos_historico.csv"))
    ap.add_argument("--export-date", default=None,
                    help="Fecha del export (YYYY-MM-DD). Por defecto: la fecha de "
                         "modificación del archivo diario.")
    args = ap.parse_args()

    if not os.path.exists(args.daily):
        sys.exit(f"No se encontró el export diario: {args.daily}")

    if args.export_date:
        export_date = datetime.strptime(args.export_date, "%Y-%m-%d").date()
    else:
        export_date = date.fromtimestamp(os.path.getmtime(args.daily))

    existing = read_csv(args.hist)
    daily = read_xlsx(args.daily, export_date)

    seen = {row_key(r) for r in existing}
    added = 0
    merged = list(existing)
    for r in daily:
        k = row_key(r)
        if k not in seen:
            seen.add(k)
            merged.append(r)
            added += 1

    with open(args.hist, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=HEADERS)
        w.writeheader()
        w.writerows(merged)

    convs = len({r["num_conversacion"] for r in merged if r["num_conversacion"]})
    print(f"Histórico actualizado: {args.hist}")
    print(f"  filas previas:   {len(existing)}")
    print(f"  filas del día:   {len(daily)}")
    print(f"  filas nuevas:    {added}")
    print(f"  filas totales:   {len(merged)}")
    print(f"  conversaciones:  {convs}")


if __name__ == "__main__":
    main()
