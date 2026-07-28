#!/usr/bin/env python3
"""
Reporte automático de PXI por correo (semanal / mensual) vía Resend.

Lee datos.xlsx (hoja Historial), calcula el PXI con la MISMA lógica del dashboard,
arma un correo HTML con marca Fertilidad Integral y lo envía con la API de Resend.

Uso:
    python scripts/send_report.py --period week
    python scripts/send_report.py --period month

Variables de entorno (se configuran como GitHub Secrets):
    RESEND_API_KEY   clave de la API de Resend
    REPORT_FROM      remitente verificado, ej. "Auditoría FI <reportes@tudominio.com>"
    REPORT_TO        destinatarios separados por coma
"""
import argparse, os, re, sys, unicodedata, json
from collections import defaultdict
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl (pip install openpyxl)")

# ── Config (igual que el dashboard) ──
WEIGHTS = {"p1": 20, "p2": 20, "p3": 15, "p4": 20, "p5": 10, "p6": 15}
# Nombres en la columna `agente` que NO son coordinadoras (cuentas de sistema /
# contactos). Se excluyen de los reportes. Comparación normalizada (sin acentos,
# minúsculas) — igual que en el dashboard.
NON_AGENTS = {"hola fertilidad", "mariana sanchez"}
BOTS = {"fi bot", "atom", "api", "agendamiento - instagram", "agendamiento - facebook"}
TOK = {
 "request": ["precio","costo","cuanto","como","cuando","donde","cual","puedo","pueden","podria",
   "informacion","info","agendar","cita","quiero","necesito","me interesa","estoy interesad",
   "disponib","horario","direccion","mandame","enviame","me puedes","una pregunta","una duda","sirve","aplica"],
 "closer": ["gracias","ok","vale","perfecto","listo","igualmente","bendiciones","de nada","claro","excelente"],
 "value": ["consulta","evaluacion","valoracion","diagnostico","plan","incluye","especialista","estudio",
   "paquete","programa","ultrasonido","antimulleriana","reserva ovarica","inseminacion","in vitro","fiv",
   "congelamiento","ovodonacion","espermatobioscopia","fragmentacion","tratamiento","hormona","laboratorio",
   "perfil","check","revision","pgt","dgp","prueba","embrion","criopreserv","biopsia","muestra","analisis",
   "seleccion","sesion","servicio","contiene","consiste","abarca","comprende","contempla"],
 "clinical": ["tasa de exito","% de exito","porcentaje de exito","probabilidad de embarazo","vas a lograr",
   "vas a quedar embarazada","te garantizo","garantizamos el embarazo","eres buen candidat","eres buena candidat",
   "tu diagnostico es"],
 "prohibited": ["relajate","no te estreses","todavia eres joven","al menos puedes","todo pasa por algo",
   "muchas personas pasan por esto","infertil","ciclo fallido","fallo el ciclo","embarazo geriatrico",
   "todo va a estar bien","todos los casos tienen solucion","es una inversion en su futuro",
   "es una inversion en tu futuro","cada mes cuenta","a su edad no puede esperar","a tu edad no puede esperar",
   "si no actua ahora","madre de alquiler","vientre de alquiler"],
 "emotion": ["anos intentando","anos buscando","llevamos anos","perdida gestacional","aborto","perdi a mi bebe",
   "perdi el embarazo","desesper","no he podido embaraz","no hemos podido embaraz","no puedo quedar embaraz",
   "me siento triste","muy frustrad","estoy agotada","cansados de","mucha angustia"],
 "frustration": ["muy molesta","muy molesto","estoy molesta","estoy molesto","enojad","indignad","pesimo servicio",
   "mal servicio","nadie me ha","llevo esperando","sigo esperando","no me han contestado","no me han respondido",
   "pesima atencion","inaceptable","decepcion"],
 "validation": ["entiendo","comprendo","es valid","por supuesto","sin compromiso","te entiendo","lo siento",
   "siento mucho","lamento","una disculpa","que valiente","estamos contigo","estamos aqui","aqui estamos"],
 "highValue": ["fiv","in vitro","pgt","dgp","ovodonacion"],
}
CHATSPEAK = ["xq","xk","pq","xfa","xfis","porfa","porfis","tmb","tb","bn","tqm","xd","salu2",
  "finde","ntp","nps","q","k","d","x","pa","grax","graxias","holaa","siii","noo","aki","ke","komo"]
GREET = ["hola","buen dia","buenos dias","buenas tardes","buenas noches","que tal","bienvenid",
  "que gusto saludar","un gusto saludar"]
POLITE = ["por favor","porfavor","gracias","con gusto","con mucho gusto","claro que si","encantad",
  "quedo atent","estoy para ayudar","estoy para servir","con todo gusto","sera un placer","que tenga",
  "excelente dia","feliz dia","feliz tarde","un gusto","para servirte","no dude","cualquier duda",
  "quedo a la orden","a la orden","quedo pendiente","permiteme","con gusto te","te comparto","te ayudo",
  "te apoyo","te puedo ayudar","quedo al pendiente","estamos al pendiente","estamos para","no te preocupes",
  "no hay problema","con todo el gusto","sera un gusto","con cariño","un abrazo","que tengas",
  "saludos cordiales","muchas gracias","mil gracias","claro que","por nada","estamos en contacto",
  # Frases cálidas adicionales sugeridas por el equipo (feedback Amabilidad).
  "gracias por compartirme","gracias por compartir","permite un momento","permiteme un momento",
  "sigo contigo","estoy contigo","con mucho cariño","no te preocupe","no se preocupe"]
# Apertura cálida por PRESENTACIÓN (alternativa al saludo cuando la asesora abre el hilo).
INTRO = ["mi nombre es","me llamo","soy tu asesora","soy tu asesor","un placer atenderte",
  "un placer poderte ayudar","un placer poder ayudarte","con quien tengo el gusto","con quien tengo el placer",
  "sera un placer atenderte","sera un gusto atenderte","un gusto atenderte","un gusto poder ayudarte"]
# Cierre cordial (solo se exige cuando la asesora tuvo el último turno).
CLOSING = ["quedo al pendiente","quedo pendiente","estamos al pendiente","estamos en contacto",
  "seguimos en contacto","quedo atent","cualquier duda","cualquier cosa","cualquier pregunta","me avisas",
  "me comentas","que tengas","que tenga","excelente dia","feliz dia","feliz tarde","buen dia","saludos",
  "estamos para servirte","estamos para ayudarte","estoy para ayudar","estoy para servir","quedo a la orden",
  "a la orden","con gusto te ayudo","no dudes en"]
# Frases frías. HARD = trato descortés → fuerza nivel Deficiente (0).
COLD_HARD = ["no me importa","no hacemos lo que quieres","no hacemos lo que quiere","yo no se",
  "no te entiendo","no le entiendo","no es mi problema","ya te dije","tienes que entender"]
# SOFT = trato seco/impersonal → limita el nivel a Aceptable (70) como máximo.
COLD_SOFT = ["esperame","espere","explicame bien","explique bien","estos son los precios",
  "ese es el precio","tienes que ver a un doctor","tiene que ver a un doctor","solo te puedo decir"]
# Detección de emoji en el texto de la asesora (bonus de calidez).
EMOJI = re.compile("[\U0001F000-\U0001FAFF\U00002600-\U000027BF\U00002B00-\U00002BFF❤♥☺✅]")
PILLAR_LABELS = {"p1":"P1 Velocidad","p2":"P2 Atención plena","p3":"P3 Valor antes de precio",
                 "p4":"P4 Lenguaje seguro","p5":"P5 Calidad de redacción","p6":"P6 Amabilidad y cortesía"}
FLAG_LABELS = {"r1":"R1 Frustración no validada","r2":"R2 Info clínica sin confirmación",
               "r3":"R3 Alto valor sin supervisor","r4":"R4 Agente >2h en horario laboral",
               "r5":"R5 Pendiente para turno matutino"}
# Brand
CREAM,IVORY,FOREST,SAGE,FSOFT,BORDER = "#F2F3E9","#F4F3E1","#20281B","#738D84","#3D4E36","#E8E7D8"
OK,WARN,BAD = "#6B9E6E","#C9A24B","#C0574A"

def norm(s):
    s = str(s or "").lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

def clean(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()

# Cómo se MUESTRA el nombre de una coordinadora (sin afectar la agrupación interna).
# Itzel se escribe siempre sin acento, por preferencia.
NAME_DISPLAY = {"itzel rodriguez": "Itzel Rodriguez"}
def display_name(n):
    c = clean(n)
    return NAME_DISPLAY.get(norm(c).strip(), c)

def has_any(text, toks):
    return any(t in text for t in toks)

def parse_hora(h):
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4}).*?(\d{1,2}):(\d{2})\s*(am|pm)?", str(h or ""), re.I)
    if not m:
        return None
    d, mo, y, hh, mm, ap = m.groups()
    hh = int(hh)
    if ap:
        ap = ap.lower()
        if ap == "pm" and hh < 12: hh += 12
        if ap == "am" and hh == 12: hh = 0
    return datetime(int(y), int(mo), int(d), hh, int(mm))

def needs_reply(text):
    t = norm(text).strip()
    if not t: return False
    if "?" in t or "¿" in str(text): return True
    # "en cuanto" = "tan pronto como": NO es solicitud de precio. Sin este guard,
    # "En cuanto realice la transferencia…" matcheaba "cuanto" y marcaba el chat
    # cerrado como pregunta pendiente (falso positivo).
    t = re.sub(r"\ben cuanto\b", " ", t)
    # Cierre/cortesía sin pregunta ("igualmente", "muchas gracias") no requiere respuesta.
    if has_any(t, TOK["closer"]) and not has_any(t, TOK["request"]): return False
    if has_any(t, TOK["request"]): return True
    return False

PRICE_RE = re.compile(r"\$\s?\d|\b\d{3,}\b\s*(pesos|mxn|mil)|cuesta|tiene un costo|el costo es|el precio es|son \$", re.I)

def _emit(groups, g, r):
    # `num_conversacion` NO es único: Atom reutiliza esos números entre
    # conversaciones distintas, así que agrupar por él mezcla pacientes y
    # genera huecos falsos. El identificador real es la URL de Atom; agrupamos
    # por URL y, si faltara, caemos al num_conversacion.
    num = str(g(r, "num_conversacion", "conversacion", "id") or "").strip()
    url = clean(g(r, "url"))
    key = url or num
    if not key:
        return
    groups[key].append({
        "tipo": norm(g(r, "tipo")), "direccion": norm(g(r, "direccion")),
        "remitente": clean(g(r, "remitente")), "contenido": str(g(r, "contenido") or ""),
        "hora": parse_hora(g(r, "hora")), "agente": clean(g(r, "agente")),
        "tipificacion": clean(g(r, "tipificacion")), "es_venta": norm(g(r, "es_venta", "es venta")),
        "url": url, "num": num,
    })


def load(path):
    """Carga conversaciones desde el histórico (.csv) o un export diario (.xlsx).

    El histórico (datos_historico.csv) ya trae fechas absolutas resueltas por
    merge_history.py, así que el reporte ve TODA la semana/mes acumulada, no un
    solo día.
    """
    groups = defaultdict(list)
    if path.lower().endswith(".csv"):
        import csv as _csv
        with open(path, newline="", encoding="utf-8") as f:
            rdr = _csv.DictReader(f)
            idx = {norm(h).strip(): h for h in (rdr.fieldnames or [])}
            def g(row, *names):
                for n in names:
                    if n in idx: return row.get(idx[n], "")
                return ""
            for row in rdr:
                _emit(groups, g, row)
        return groups
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Historial"] if "Historial" in wb.sheetnames else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [norm(h).strip() for h in next(it)]
    idx = {h: i for i, h in enumerate(hdr)}
    def g(row, *names):
        for n in names:
            if n in idx: return row[idx[n]]
        return ""
    for r in it:
        _emit(groups, g, r)
    return groups

# Working hours = 07:00–24:00; off-hours = 00:00–07:00. SLA clock pauses off-hours.
WORK_START_H = 7  # working hours run 7 a.m.–midnight (full coverage via shifts)
def working_minutes_between(t0, t1):
    if not t0 or not t1 or t1 <= t0:
        return 0
    from datetime import timedelta
    total = 0.0
    cur = t0
    while cur < t1:
        day_start = cur.replace(hour=WORK_START_H, minute=0, second=0, microsecond=0)
        next_midnight = (cur.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1))
        win_start = max(cur, day_start)
        win_end = min(t1, next_midnight)
        if win_end > win_start:
            total += (win_end - win_start).total_seconds()
        cur = next_midnight
    return round(total / 60)

def score(num, msgs):
    chat = sorted([m for m in msgs if m["tipo"] == "mensaje"], key=lambda m: m["hora"] or datetime.min)
    inbound = [m for m in chat if m["direccion"] == "entrante"]
    out = [m for m in chat if m["direccion"] == "saliente" and m["remitente"].lower() not in BOTS]
    if not inbound or not out:
        return None
    agente = clean(next((m["agente"] for m in msgs if clean(m["agente"])), "")) or \
             (max({m["remitente"]: sum(1 for x in out if x["remitente"] == m["remitente"]) for m in out}.items(),
                  key=lambda kv: kv[1])[0] if out else "") or "Sin asignar"
    agente = display_name(agente)
    atext = norm("  ".join(m["contenido"] for m in out))
    ptext = norm("  ".join(m["contenido"] for m in inbound))
    pillars = {}

    # P1 Speed (working hours 7am–midnight; clock pauses overnight).
    # Two-tier SLA: 1st reply < 2 min; every later reply within 15 min.
    # El reloj arranca en el HANDOFF (cuando Atom asigna el chat a la asesora), no en el
    # primer mensaje del lead: el bot atiende la calificación inicial al instante y el
    # enrutamiento no es responsabilidad de la asesora. Sin evento de asignación, caemos
    # al primer mensaje entrante.
    first_in = inbound[0]["hora"]
    first_reply = next((m["hora"] for m in out if m["hora"] and first_in and m["hora"] >= first_in), None)
    assign_ts = [m["hora"] for m in msgs
                 if m["tipo"] == "evento" and m["hora"]
                 and re.search(r"asignad[oa]", m["contenido"] or "", re.I)
                 and not re.search(r"al bot", m["contenido"] or "", re.I)]
    sla_start = first_in
    if first_reply and first_in:
        handoffs = [t for t in assign_ts if t and t <= first_reply]
        if handoffs:
            sla_start = max([first_in] + handoffs)
    frm = working_minutes_between(sla_start, first_reply) if (sla_start and first_reply) else None

    max_sub = None
    p_since = None
    seen_first = False
    for m in chat:
        if not m["hora"]:
            continue
        if m["direccion"] == "entrante":
            if p_since is None and needs_reply(m["contenido"]):
                p_since = m["hora"]
        elif m["direccion"] == "saliente":
            if not seen_first:
                seen_first = True; p_since = None; continue
            if p_since is not None:
                wm = working_minutes_between(p_since, m["hora"])
                max_sub = wm if max_sub is None else max(max_sub, wm)
                p_since = None

    first_ok = frm is not None and frm <= 2
    sub_ok = max_sub is None or max_sub <= 15
    pillars["p1"] = {"applies": frm is not None, "pass": (first_ok and sub_ok)}

    last = chat[-1]
    dropped = last["direccion"] == "entrante" and needs_reply(last["contenido"])
    pillars["p2"] = {"applies": True, "pass": not dropped}

    p3_applies = p3_pass = None
    gave_price = False
    for i, m in enumerate(out):
        if PRICE_RE.search(norm(m["contenido"])) or re.search(r"\$\s?\d", m["contenido"]):
            gave_price = True; p3_applies = True
            before = norm(" ".join(x["contenido"] for x in out[:i+1]))
            after = norm(" ".join(x["contenido"] for x in out[i+1:i+3]))
            if has_any(before, TOK["value"]) or has_any(after, TOK["value"]):
                p3_pass = True
            elif p3_pass is None:
                p3_pass = False
    pillars["p3"] = {"applies": bool(p3_applies), "pass": p3_pass if p3_applies else None}

    clinical = False
    for claim in TOK["clinical"]:
        i = atext.find(claim)
        while i != -1:
            pre = atext[max(0, i-12):i]
            if not re.search(r"\b(no|sin)\s$", pre): clinical = True; break
            i = atext.find(claim, i+1)
        if clinical: break
    prohibited = has_any(atext, TOK["prohibited"])
    pillars["p4"] = {"applies": True, "pass": not (clinical or prohibited)}

    frus = has_any(ptext, TOK["frustration"])
    validated = has_any(atext, TOK["validation"])

    # P5 Writing Quality — style heuristic on the agent's own substantial messages.
    substantial = [m for m in out
                   if len(re.sub(r"\s+", " ", (m["contenido"] or "").strip()).split(" ")) >= 3
                   and len(re.sub(r"[^a-záéíóúñ]", "", m["contenido"] or "", flags=re.I)) >= 12]
    bad = 0
    for m in substantial:
        raw = (m["contenido"] or "").strip()
        letters = re.sub(r"[^a-záéíóúñ]", "", raw, flags=re.I)
        issue = False
        fl = re.search(r"[a-záéíóúñ]", raw, re.I)
        if fl and fl.group(0).islower(): issue = True
        tok = " " + re.sub(r"\s+", " ", re.sub(r"[^a-z0-9ñ ]", " ", norm(raw), flags=re.I)).strip() + " "
        if any((" " + w + " ") in tok for w in CHATSPEAK): issue = True
        if len(letters) >= 8:
            ups = len(re.findall(r"[A-ZÁÉÍÓÚÑ]", raw))
            if ups / len(letters) > 0.8: issue = True
        if "?" in raw and "¿" not in raw: issue = True
        if issue: bad += 1
    pillars["p5"] = ({"applies": True, "pass": (bad / len(substantial)) <= 0.2}
                     if substantial else {"applies": False, "pass": None})

    # P6 Amabilidad y cortesía — rúbrica por niveles (Excelente=100/Aceptable=70/Deficiente=0).
    # Consciente del contexto (feedback del equipo):
    #  · Saludo/presentación solo se exige si la asesora ABRE el hilo.
    #  · Cierre cordial solo se exige si la asesora tuvo el ÚLTIMO turno.
    #  · Frases frías restan; calidez/empatía/emoji suman. No se penaliza la brevedad.
    first_out = norm(out[0]["contenido"]) if out else ""
    last_out = norm(out[-1]["contenido"]) if out else ""
    agent_opened = bool(out) and chat.index(out[0]) <= 1
    agent_had_last = bool(chat) and chat[-1]["direccion"] == "saliente"
    hard_cold = has_any(atext, COLD_HARD)
    soft_cold = has_any(atext, COLD_SOFT)
    warm_open = (has_any(first_out, GREET) or has_any(first_out, INTRO)) if agent_opened else True
    courtesy = has_any(atext, POLITE) or validated
    proper_close = (has_any(last_out, CLOSING) or has_any(last_out, POLITE)) if agent_had_last else True
    warm_bonus = bool(EMOJI.search(atext))
    if hard_cold:
        p6_score = 0
    else:
        missed = (0 if warm_open else 1) + (0 if courtesy else 1) + (0 if proper_close else 1)
        if missed == 0:
            p6_score = 100
        elif missed == 1:
            p6_score = 100 if warm_bonus else 70
        else:
            p6_score = 70 if warm_bonus else 0
        if soft_cold:
            p6_score = min(p6_score, 70)
    pillars["p6"] = {"applies": True, "pass": p6_score >= 70, "score": p6_score}

    alltext = norm(" ".join(m["contenido"] for m in chat))
    # Response-time analysis (misma lógica afinada que el dashboard):
    #  · Solo cuentan retrasos del MISMO día. Un hueco que cruza al día siguiente
    #    = lead frío / re-enganche, NO un incumplimiento >2h.
    #  · Mensajes nocturnos (12am–7am) entran a la cola matutina: deben atenderse
    #    antes de las 8am; si se responden después, es tardío.
    #  · Pregunta en horario laboral sin responder al final del snapshot NO se
    #    marca como >2h (lo captura P2).
    max_wait_working = 0
    agent_slow = False
    morning_queue = False
    pending_since = None
    for m in chat:
        if not m["hora"]:
            continue
        if m["direccion"] == "entrante":
            if pending_since is None and needs_reply(m["contenido"]):
                pending_since = m["hora"]
        elif m["direccion"] == "saliente" and pending_since is not None:
            reply = m["hora"]
            if pending_since.hour < WORK_START_H:
                max_wait_working = max(max_wait_working, working_minutes_between(pending_since, reply))
                deadline = pending_since.replace(hour=WORK_START_H + 1, minute=0, second=0, microsecond=0)
                if reply > deadline:
                    agent_slow = True
            elif pending_since.date() == reply.date():
                wm = working_minutes_between(pending_since, reply)
                max_wait_working = max(max_wait_working, wm)
                if wm > 120:
                    agent_slow = True
            # else: respondió otro día → sin incumplimiento.
            pending_since = None
    if pending_since is not None:  # thread ends with the lead still waiting
        if pending_since.hour < WORK_START_H:
            morning_queue = True   # llegó de noche → cola matutina (R5)
        # En horario laboral sin responder al cierre: lo captura P2, no R4/R5.
    flags = {
        "r1": frus and not validated,
        "r2": clinical,
        "r3": gave_price and has_any(alltext, TOK["highValue"]),
        "r4": agent_slow,
        "r5": morning_queue,
    }
    ts = next((m["hora"] for m in chat if m["hora"]), None)
    es_venta = next((m["es_venta"] for m in msgs if m["es_venta"]), "")
    url = next((m["url"] for m in msgs if m["url"]), "")
    num = next((m["num"] for m in msgs if m.get("num")), num)  # nº real (la clave puede ser la URL)
    return {"num": num, "agente": agente, "pillars": pillars, "flags": flags, "frm": frm,
            "max_wait_working": max_wait_working, "ts": ts, "es_venta": es_venta, "url": url}

def pillar_val(p):
    # Valor numérico (0..100) de un pilar: los binarios usan pass→100/0; los
    # graduados (p. ej. P6) llevan un `score` (0/70/100).
    return p["score"] if p.get("score") is not None else (100 if p["pass"] else 0)

def pxi_of(convs):
    w = s = 0
    pill = {}
    for k in WEIGHTS:
        appl = [c for c in convs if c["pillars"][k]["applies"]]
        pill[k] = round(sum(pillar_val(c["pillars"][k]) for c in appl) / len(appl)) if appl else None
        if pill[k] is not None:
            w += WEIGHTS[k]; s += WEIGHTS[k] * pill[k]
    return (round(s / w) if w else None), pill

# Consejo accionable por pilar (qué hacer para subir esa categoría la próxima semana).
PILLAR_TIP = {
    "p1": "Responder el primer mensaje en menos de 2 min y no dejar pasar más de 15 min entre respuestas.",
    "p2": "No cerrar el turno con un mensaje del paciente sin responder; revisar hilos abiertos antes de salir.",
    "p3": "Antes de dar un precio, explicar el valor: qué incluye, evaluación, especialista, acompañamiento.",
    "p4": "Evitar promesas clínicas o frases minimizadoras; usar lenguaje prudente ('los resultados varían…').",
    "p5": "Cuidar la redacción: mayúscula inicial, abrir con ¿, sin abreviaturas informales ni MAYÚSCULAS.",
    "p6": "Abrir con saludo/presentación (si abres el chat), usar frases cálidas y empáticas, "
          "y cerrar con cortesía cuando das el último mensaje ('con gusto', 'quedo al pendiente'). "
          "Evita frases secas ('espérame', 'estos son los precios').",
}

def coaching_focus(convs):
    """Para una coordinadora: devuelve (pilar_más_débil, pct, conversación_ejemplo).

    El pilar más débil = la categoría aplicable con menor % de cumplimiento.
    El ejemplo = una conversación donde ese pilar aplica y NO se cumplió
    (de preferencia con URL para abrirla).
    """
    _, pill = pxi_of(convs)
    applicable = {k: v for k, v in pill.items() if v is not None}
    if not applicable:
        return None, None, None
    weakest = min(applicable, key=lambda k: applicable[k])
    fails = [c for c in convs
             if c["pillars"][weakest]["applies"] and not c["pillars"][weakest]["pass"]]
    fails.sort(key=lambda c: (0 if c["url"] else 1, -(c["ts"].timestamp() if c["ts"] else 0)))
    example = fails[0] if fails else None
    return weakest, applicable[weakest], example

def period_range(period, today=None, offset=1):
    # offset = cuántos periodos hacia atrás: 1 = el periodo COMPLETO anterior
    # (comportamiento programado por defecto), 0 = el periodo ACTUAL en curso.
    today = today or datetime.now()
    if period == "week":
        this_mon = (today - timedelta(days=today.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        start = this_mon - timedelta(days=7 * offset)
        end = start + timedelta(days=7)
        label = f"semana del {start.strftime('%d/%m/%Y')}"
    else:
        first_this = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        start = first_this
        for _ in range(offset):
            start = (start - timedelta(days=1)).replace(day=1)
        end = (start + timedelta(days=32)).replace(day=1)
        meses = ["enero","febrero","marzo","abril","mayo","junio","julio","agosto",
                 "septiembre","octubre","noviembre","diciembre"]
        label = f"{meses[start.month-1]} {start.year}"
    return start, end, label

def color_for(v):
    if v is None: return FSOFT
    return OK if v >= 80 else WARN if v >= 60 else BAD

def build_html(period_label, cur, prev, errors):
    team_pxi, _ = pxi_of([c for a in cur.values() for c in a])
    prev_team, _ = pxi_of([c for a in prev.values() for c in a]) if prev else (None, None)
    all_cur = [c for a in cur.values() for c in a]
    sla_appl = [c for c in all_cur if c["pillars"]["p1"]["applies"]]
    sla = round(sum(1 for c in sla_appl if c["pillars"]["p1"]["pass"]) / len(sla_appl) * 100) if sla_appl else None
    citas = sum(1 for c in all_cur if c["es_venta"] == "si")
    n_flags = sum(sum(1 for v in c["flags"].values() if v) for c in all_cur)

    def kpi(label, val):
        return (f'<td style="padding:14px;background:{IVORY};border:1px solid {BORDER};border-radius:8px;'
                f'text-align:center"><div style="font-size:12px;color:{FSOFT};text-transform:uppercase;'
                f'letter-spacing:.5px">{label}</div><div style="font-size:30px;color:{FOREST};font-weight:bold;'
                f'margin-top:4px">{val}</div></td>')

    delta_team = f" ({'+' if (team_pxi or 0)-(prev_team or 0)>=0 else ''}{(team_pxi or 0)-(prev_team or 0)} vs. periodo previo)" if (team_pxi is not None and prev_team is not None) else ""

    rows = ""
    for agent in sorted(cur, key=lambda a: -(pxi_of(cur[a])[0] or -1)):
        p, pill = pxi_of(cur[agent])
        pp, _ = pxi_of(prev.get(agent, [])) if prev else (None, None)
        if pp is not None and p is not None:
            d = p - pp
            trend = (f'<span style="color:{OK}">▲ +{d}</span>' if d >= 3 else
                     f'<span style="color:{BAD}">▼ {d}</span>' if d <= -3 else
                     f'<span style="color:{SAGE}">▬ {d:+d}</span>')
        else:
            trend = f'<span style="color:{FSOFT}">—</span>'
        cells = "".join(
            f'<td style="text-align:center;padding:8px;color:{color_for(pill[k])};font-weight:bold">'
            f'{pill[k] if pill[k] is not None else "N/A"}</td>' for k in WEIGHTS)
        rows += (f'<tr style="border-bottom:1px solid {BORDER}">'
                 f'<td style="padding:8px 10px;font-weight:bold;color:{FOREST}">{agent}</td>'
                 f'<td style="text-align:center;padding:8px;font-weight:bold;font-size:18px;color:{color_for(p)}">{p if p is not None else "—"}</td>'
                 f'<td style="text-align:center;padding:8px">{trend}</td>'
                 f'{cells}'
                 f'<td style="text-align:center;padding:8px;color:{FSOFT}">{len(cur[agent])}</td></tr>')

    flag_rows = ""
    for c in errors:
        for f, on in c["flags"].items():
            if on:
                link = f'<a href="{c["url"]}" style="color:{SAGE}">Abrir ↗</a>' if c["url"] else ""
                flag_rows += (f'<tr style="border-bottom:1px solid {BORDER}">'
                              f'<td style="padding:7px 10px;color:{BAD};font-weight:bold;white-space:nowrap">{FLAG_LABELS[f]}</td>'
                              f'<td style="padding:7px 10px;color:{FOREST}">{c["agente"]}</td>'
                              f'<td style="padding:7px 10px;color:{FSOFT}">conv. {c["num"]}</td>'
                              f'<td style="padding:7px 10px">{link}</td></tr>')
    if not flag_rows:
        flag_rows = f'<tr><td colspan="4" style="padding:14px;color:{OK};text-align:center">✓ Sin incidencias en el periodo</td></tr>'

    # Foco de coaching: por coordinadora, su pilar más débil + una conversación de ejemplo.
    coach_rows = ""
    for agent in sorted(cur, key=lambda a: (pxi_of(cur[a])[0] if pxi_of(cur[a])[0] is not None else 101)):
        wk, pct, ex = coaching_focus(cur[agent])
        if wk is None:
            continue
        link = (f'<a href="{ex["url"]}" style="color:{SAGE}">Abrir conv. {ex["num"]} ↗</a>'
                if (ex and ex["url"]) else (f'conv. {ex["num"]}' if ex else "—"))
        coach_rows += (f'<tr style="border-bottom:1px solid {BORDER}">'
                       f'<td style="padding:9px 10px;font-weight:bold;color:{FOREST};white-space:nowrap;vertical-align:top">{agent}</td>'
                       f'<td style="padding:9px 10px;vertical-align:top;white-space:nowrap">'
                       f'<span style="color:{color_for(pct)};font-weight:bold">{PILLAR_LABELS[wk]}</span>'
                       f'<div style="font-size:11px;color:{FSOFT}">{pct}% cumplimiento</div></td>'
                       f'<td style="padding:9px 10px;color:{FSOFT};vertical-align:top">{PILLAR_TIP[wk]}'
                       f'<div style="font-size:11px;margin-top:4px">{link}</div></td></tr>')
    if not coach_rows:
        coach_rows = f'<tr><td colspan="3" style="padding:14px;color:{OK};text-align:center">✓ Sin datos suficientes para coaching</td></tr>'

    pill_head = "".join(f'<th style="padding:8px;font-size:11px;color:{FSOFT}">{PILLAR_LABELS[k].split()[0]}</th>' for k in WEIGHTS)

    return f"""<!DOCTYPE html><html><body style="margin:0;background:{CREAM};font-family:'DM Sans',Arial,sans-serif;color:{FOREST}">
<div style="max-width:760px;margin:0 auto;padding:24px">
  <div style="background:{FOREST};border-radius:10px;padding:28px 24px;color:{CREAM}">
    <div style="font-size:13px;letter-spacing:1px;color:{SAGE}">FERTILIDAD INTEGRAL · AUDITORÍA DE VENTAS</div>
    <div style="font-size:26px;font-weight:bold;margin-top:6px">Reporte PXI — {period_label}</div>
  </div>

  <h2 style="font-size:16px;margin:26px 0 12px">Resumen del equipo</h2>
  <table width="100%" cellspacing="8" cellpadding="0"><tr>
    {kpi("PXI clínica", (str(team_pxi) if team_pxi is not None else "—"))}
    {kpi("Conversaciones", len(all_cur))}
    {kpi("Cumple SLA velocidad", (f"{sla}%" if sla is not None else "—"))}
    {kpi("Citas", citas)}
    {kpi("Alertas", n_flags)}
  </tr></table>
  <p style="font-size:13px;color:{FSOFT};margin:8px 2px 0">PXI del equipo: <b>{team_pxi if team_pxi is not None else "—"}</b>{delta_team}</p>

  <h2 style="font-size:16px;margin:28px 0 12px">Desempeño por coordinadora</h2>
  <table width="100%" style="border-collapse:collapse;background:{IVORY};border:1px solid {BORDER};border-radius:8px;font-size:13px">
    <thead><tr style="background:{CREAM}">
      <th style="text-align:left;padding:8px 10px;font-size:11px;color:{FSOFT}">Coordinadora</th>
      <th style="padding:8px;font-size:11px;color:{FSOFT}">PXI</th>
      <th style="padding:8px;font-size:11px;color:{FSOFT}">Tendencia</th>
      {pill_head}
      <th style="padding:8px;font-size:11px;color:{FSOFT}">Convs</th>
    </tr></thead><tbody>{rows}</tbody>
  </table>

  <h2 style="font-size:16px;margin:28px 0 12px">Foco de coaching por coordinadora</h2>
  <table width="100%" style="border-collapse:collapse;background:{IVORY};border:1px solid {BORDER};border-radius:8px;font-size:13px">
    <thead><tr style="background:{CREAM}">
      <th style="text-align:left;padding:8px 10px;font-size:11px;color:{FSOFT}">Coordinadora</th>
      <th style="text-align:left;padding:8px 10px;font-size:11px;color:{FSOFT}">Pilar más débil</th>
      <th style="text-align:left;padding:8px 10px;font-size:11px;color:{FSOFT}">Acción sugerida + ejemplo</th>
    </tr></thead><tbody>{coach_rows}</tbody>
  </table>

  <h2 style="font-size:16px;margin:28px 0 12px">Incidencias del periodo</h2>
  <table width="100%" style="border-collapse:collapse;background:{IVORY};border:1px solid {BORDER};border-radius:8px;font-size:13px">
    <tbody>{flag_rows}</tbody>
  </table>

  <p style="font-size:11px;color:{FSOFT};margin-top:28px;line-height:1.6">
    Generado automáticamente desde el dashboard de auditoría. El export de Atom es solo texto:
    los puntos de muestreo manual (uso del nombre, SPIN, objeciones, HubSpot) no se incluyen aquí.
    Ver dashboard completo: https://annika-fertilidad.github.io/Auditoria-Equipo-Ventas/
  </p>
</div></body></html>"""

def send(subject, html):
    import urllib.request, urllib.error
    key = os.environ.get("RESEND_API_KEY", "")
    sender = os.environ.get("REPORT_FROM", "")
    to = [x.strip() for x in os.environ.get("REPORT_TO", "").split(",") if x.strip()]

    # Diagnóstico claro de configuración (sin revelar la clave)
    print(f"Remitente (REPORT_FROM): {sender!r}")
    print(f"Destinatarios (REPORT_TO): {to}")
    print(f"Clave (RESEND_API_KEY): {'definida, empieza con ' + key[:4] + '...' if key else 'NO DEFINIDA'}")
    problemas = []
    if not key:
        problemas.append("Falta el secreto RESEND_API_KEY.")
    elif not key.startswith("re_"):
        problemas.append("RESEND_API_KEY no empieza con 're_' (¿está mal copiada?).")
    if not sender:
        problemas.append("Falta el secreto REPORT_FROM.")
    if not to:
        problemas.append("Falta el secreto REPORT_TO (o está vacío).")
    if problemas:
        print("PROBLEMAS DE CONFIGURACIÓN:")
        for p in problemas:
            print("  -", p)

    payload = json.dumps({"from": sender, "to": to, "subject": subject, "html": html}).encode()
    req = urllib.request.Request("https://api.resend.com/emails", data=payload, method="POST",
        headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json",
                 "User-Agent": "AuditoriaFI/1.0 (+https://annika-fertilidad.github.io/Auditoria-Equipo-Ventas/)",
                 "Accept": "application/json"})
    try:
        with urllib.request.urlopen(req) as r:
            print("Enviado:", r.status, r.read().decode()[:300])
    except urllib.error.HTTPError as e:
        body = e.read().decode(errors="replace")
        print(f"ERROR de Resend (HTTP {e.code}): {body}")
        if e.code == 403:
            print("\nCausa típica del 403:")
            print("  1) La RESEND_API_KEY es inválida o no tiene permiso de envío.")
            print("  2) El correo/dominio de REPORT_FROM no está verificado en Resend.")
            print("     -> Para empezar, usa el remitente de prueba: onboarding@resend.dev")
        raise SystemExit(1)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--period", choices=["week", "month"], required=True)
    ap.add_argument("--offset", type=int, default=1,
                    help="Periodos hacia atrás: 1 = anterior completo (default), 0 = actual en curso.")
    ap.add_argument("--file", default="datos_historico.csv")
    ap.add_argument("--dry-run", action="store_true", help="No envía; escribe preview_<period>.html")
    args = ap.parse_args()

    groups = load(args.file)
    convs = [s for s in (score(n, m) for n, m in groups.items()) if s]
    # Excluir cuentas de sistema / contactos que no son coordinadoras de ventas.
    convs = [c for c in convs if norm(c["agente"]) not in NON_AGENTS]

    start, end, label = period_range(args.period, offset=args.offset)
    pstart = start - (end - start)  # previous period of same length
    def bucket(s, e):
        d = defaultdict(list)
        for c in convs:
            if c["ts"] and s <= c["ts"] < e:
                d[c["agente"]].append(c)
        return d
    cur = bucket(start, end)
    prev = bucket(pstart, start)
    errors = [c for a in cur.values() for c in a if any(c["flags"].values())]

    if not cur:
        print(f"Sin conversaciones en el periodo ({label}). No se envía correo.")
        return

    html = build_html(label, cur, prev, errors)
    subject = f"Reporte PXI {'semanal' if args.period=='week' else 'mensual'} — {label}"
    if args.dry_run:
        out = f"preview_{args.period}.html"
        open(out, "w", encoding="utf-8").write(html)
        print("Preview escrito:", out)
    else:
        send(subject, html)

if __name__ == "__main__":
    main()
